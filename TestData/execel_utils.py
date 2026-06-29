import openpyxl


def get_row_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    sh = workbook[sheet]
    return sh.max_row


def get_column_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    sh = workbook[sheet]
    return sh.max_column


def read_data(file, sheet, row, column):
    workbook = openpyxl.load_workbook(file)
    sh = workbook[sheet]
    return sh.cell(row, column).value


def write_data(file, sheet, row, column, data):
    workbook = openpyxl.load_workbook(file)
    sh = workbook[sheet]
    sh.cell(row, column).value = data
    workbook.save(file)